import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import openpyxl


class ExcelWrapperApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Translation Word Wrapper")
        self.root.geometry("450x380")
        self.root.resizable(False, False)
        
        self.filepath = tk.StringVar()
        self.preserve_symbols = tk.BooleanVar(value=True)
        
        self.create_widgets()

    def create_widgets(self):
        padding = {'padx': 10, 'pady': 5}
        
        # --- File Selection ---
        frame_file = ttk.LabelFrame(self.root, text="1. Select Excel File")
        frame_file.pack(fill="x", padx=10, pady=10)
        
        ttk.Button(frame_file, text="Browse...", command=self.browse_file).pack(side="left", **padding)
        ttk.Label(frame_file, textvariable=self.filepath, foreground="blue").pack(side="left", fill="x", expand=True, **padding)

        # --- Settings ---
        frame_settings = ttk.LabelFrame(self.root, text="2. Wrapping Settings")
        frame_settings.pack(fill="x", padx=10, pady=5)

        # Start Row
        ttk.Label(frame_settings, text="Start Row:").grid(row=0, column=0, sticky="e", **padding)
        self.entry_start_row = ttk.Entry(frame_settings, width=10)
        self.entry_start_row.insert(0, "1")
        self.entry_start_row.grid(row=0, column=1, sticky="w", **padding)

        # End Row
        ttk.Label(frame_settings, text="End Row:").grid(row=0, column=2, sticky="e", **padding)
        self.entry_end_row = ttk.Entry(frame_settings, width=10)
        self.entry_end_row.grid(row=0, column=3, sticky="w", **padding)
        ttk.Label(frame_settings, text="(Leave blank for all)").grid(row=0, column=4, sticky="w")

        # Char Count
        ttk.Label(frame_settings, text="Max Chars per Line:").grid(row=1, column=0, sticky="e", **padding)
        self.entry_chars = ttk.Entry(frame_settings, width=10)
        self.entry_chars.insert(0, "40")
        self.entry_chars.grid(row=1, column=1, sticky="w", **padding)

        # Preserve Options
        ttk.Label(frame_settings, text="Existing '◙':").grid(row=2, column=0, sticky="e", **padding)
        frame_radio = ttk.Frame(frame_settings)
        frame_radio.grid(row=2, column=1, columnspan=3, sticky="w", **padding)
        ttk.Radiobutton(frame_radio, text="Preserve", variable=self.preserve_symbols, value=True).pack(side="left", padx=(0, 10))
        ttk.Radiobutton(frame_radio, text="Remove", variable=self.preserve_symbols, value=False).pack(side="left")

        # --- Process Button ---
        frame_action = ttk.Frame(self.root)
        frame_action.pack(fill="x", padx=10, pady=20)
        
        self.btn_process = ttk.Button(frame_action, text="Process Excel File", command=self.process_file)
        self.btn_process.pack(fill="x", ipady=5)
        
        self.status_label = ttk.Label(self.root, text="Ready.", foreground="grey")
        self.status_label.pack(side="bottom", pady=5)

    def browse_file(self):
        filepath = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
        )
        if filepath:
            self.filepath.set(filepath)

    def wrap_text(self, text, max_chars, preserve):
        if text is None:
            return None
        
        text = str(text)

        if not preserve:
            # Replace existing symbols with space, then normalize multiple spaces
            text = text.replace('◙', ' ')
            text = re.sub(r'\s+', ' ', text).strip()
            paragraphs = [text]
        else:
            # Split by existing ◙ to preserve hard breaks
            paragraphs = text.split('◙')

        wrapped_paragraphs = []
        for p in paragraphs:
            words = [w for w in p.split(' ') if w] # Split by space, ignore empty
            if not words:
                wrapped_paragraphs.append("")
                continue

            lines = []
            current_line = []
            current_len = 0

            for word in words:
                space_len = 1 if current_line else 0
                if current_len + len(word) + space_len <= max_chars:
                    current_line.append(word)
                    current_len += len(word) + space_len
                else:
                    if current_line:
                        lines.append(" ".join(current_line))
                    current_line = [word]
                    current_len = len(word)

            if current_line:
                lines.append(" ".join(current_line))

            wrapped_paragraphs.append("◙".join(lines))

        return "◙".join(wrapped_paragraphs)

    def process_file(self):
        file_path = self.filepath.get()
        if not file_path or not os.path.exists(file_path):
            messagebox.showerror("Error", "Please select a valid Excel file first.")
            return

        try:
            start_row = int(self.entry_start_row.get())
            max_chars = int(self.entry_chars.get())
        except ValueError:
            messagebox.showerror("Error", "Start Row and Max Chars must be numbers.")
            return

        end_row_str = self.entry_end_row.get().strip()
        
        self.status_label.config(text="Processing...", foreground="blue")
        self.root.update()

        try:
            # 1. Open workbook
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

            # Determine end row
            end_row = int(end_row_str) if end_row_str else sheet.max_row
            preserve = self.preserve_symbols.get()

            # 2. Iterate and wrap Column C (Index 3)
            changes_made = 0
            for row_idx in range(start_row, end_row + 1):
                cell = sheet.cell(row=row_idx, column=3) # Column C
                if cell.value:
                    new_text = self.wrap_text(cell.value, max_chars, preserve)
                    if new_text != cell.value:
                        cell.value = new_text
                        changes_made += 1

            # 3. Rename original file to .bac
            backup_path = file_path + ".bac"
            # os.replace handles overwriting if a .bac already exists
            os.replace(file_path, backup_path)

            # 4. Save new file as original name
            wb.save(file_path)

            self.status_label.config(text="Done!", foreground="green")
            messagebox.showinfo(
                "Success", 
                f"Successfully processed {changes_made} cells.\n\n"
                f"Original file backed up as:\n{os.path.basename(backup_path)}"
            )

        except PermissionError:
            self.status_label.config(text="Error: File is in use.", foreground="red")
            messagebox.showerror("Permission Error", "The Excel file is currently open in another program. Please close it and try again.")
        except Exception as e:
            self.status_label.config(text="An error occurred.", foreground="red")
            messagebox.showerror("Error", f"An unexpected error occurred:\n{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelWrapperApp(root)
    root.mainloop()